import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from datetime import datetime


def export_attendance_to_excel(students, attendance_records, start_date, end_date, classroom_name):
    """
    Export attendance data to Excel file
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6b21a8", end_color="6b21a8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    present_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    absent_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    late_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    excused_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Attendance Report - {classroom_name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Date range
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"Period: {start_date} to {end_date}"
    date_cell.alignment = Alignment(horizontal="center")
    
    # Generated on
    ws.merge_cells('A3:F3')
    generated_cell = ws['A3']
    generated_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    generated_cell.alignment = Alignment(horizontal="center")

    # Headers (start from row 5)
    headers = ['Student Name', 'Adm No', 'Total Days', 'Present', 'Absent', 'Late', 'Excused', 'Attendance %']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Add data
    row = 6
    for student in students:
        records = attendance_records.get(student.id, [])
        
        total = len(records)
        present = sum(1 for r in records if r.status == 'Present')
        absent = sum(1 for r in records if r.status == 'Absent')
        late = sum(1 for r in records if r.status == 'Late')
        excused = sum(1 for r in records if r.status == 'Excused')
        
        attendance_percent = round((present / total * 100) if total > 0 else 0)
        
        ws.cell(row=row, column=1, value=f"{student.first_name} {student.last_name}").border = border
        ws.cell(row=row, column=2, value=student.registration_number).border = border
        ws.cell(row=row, column=3, value=total).border = border
        ws.cell(row=row, column=4, value=present).border = border
        ws.cell(row=row, column=5, value=absent).border = border
        ws.cell(row=row, column=6, value=late).border = border
        ws.cell(row=row, column=7, value=excused).border = border
        
        percent_cell = ws.cell(row=row, column=8, value=f"{attendance_percent}%")
        percent_cell.border = border
        
        # Color code attendance percentage
        if attendance_percent >= 80:
            percent_cell.fill = present_fill
        elif attendance_percent >= 60:
            percent_cell.fill = late_fill
        else:
            percent_cell.fill = absent_fill
        
        row += 1

    # Add summary row
    summary_row = row + 1
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
    
    # Calculate overall stats
    total_present = sum(sum(1 for r in attendance_records.get(s.id, []) if r.status == 'Present') for s in students)
    total_absent = sum(sum(1 for r in attendance_records.get(s.id, []) if r.status == 'Absent') for s in students)
    total_late = sum(sum(1 for r in attendance_records.get(s.id, []) if r.status == 'Late') for s in students)
    total_excused = sum(sum(1 for r in attendance_records.get(s.id, []) if r.status == 'Excused') for s in students)
    total_days = total_present + total_absent + total_late + total_excused
    
    ws.cell(row=summary_row, column=3, value="Total Days:").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=total_days)
    ws.cell(row=summary_row+1, column=3, value="Present:").font = Font(bold=True)
    ws.cell(row=summary_row+1, column=4, value=f"{total_present} ({round(total_present/total_days*100,1) if total_days>0 else 0}%)")
    ws.cell(row=summary_row+2, column=3, value="Absent:").font = Font(bold=True)
    ws.cell(row=summary_row+2, column=4, value=f"{total_absent} ({round(total_absent/total_days*100,1) if total_days>0 else 0}%)")
    ws.cell(row=summary_row+3, column=3, value="Late:").font = Font(bold=True)
    ws.cell(row=summary_row+3, column=4, value=f"{total_late} ({round(total_late/total_days*100,1) if total_days>0 else 0}%)")
    ws.cell(row=summary_row+4, column=3, value="Excused:").font = Font(bold=True)
    ws.cell(row=summary_row+4, column=4, value=f"{total_excused} ({round(total_excused/total_days*100,1) if total_days>0 else 0}%)")

    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    ws.column_dimensions['A'].width = 25

    # Create response
    filename = f"attendance_{classroom_name}_{start_date}_to_{end_date}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response